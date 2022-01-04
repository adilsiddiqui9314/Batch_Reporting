from tkinter import*
import tkinter as tk 
import os
#import webbrowser as wb
from webbrowser import open_new
from tkinter.messagebox import showerror, showinfo
from tkinter import font
from tkcalendar import DateEntry
import pyodbc
from openpyxl import load_workbook
from time import sleep
from win32com import client
from datetime import datetime,timedelta
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows

root = Tk()
root.title("REPORT VIEWER")
x_position = 100
y_position = 80
root.geometry(f"1150x550+{x_position}+{y_position}")
root.resizable(width=False, height=False)
root.wm_attributes("-transparentcolor", 'grey')

try:
    os.system('TASKKILL /F /IM excel.exe')

except Exception:
    print("KU")
conn = pyodbc.connect(
    'DRIVER={SQL Server};'
    'SERVER=IMRANM-LENOVO\WINCC;'
    #'SERVER='+sql_database_name+';'
    'DATABASE=DATALOG;'
    'Trusted_Connection=yes;'
) 

cursor = conn.cursor()

def print_interval_selc():
    global print_interval_type
    global get_print_interval_time
    print_interval_type = print_interval.get()
    if print_interval_type == "interval":        
        spin_interval.place(x = 900, y = 250)
        interval_label.place(x = 1000, y = 270)
        get_print_interval_time = spin_interval.get() 
        print(get_print_interval_time)
    if print_interval_type == "no_interval":
        spin_interval.place_forget()
        interval_label.place_forget()

def selection_of_report():
    global report_type
    global from_date_time
    global to_date_time
    global from_date_time1
    global to_date_time1
    report_type = report_types_system.get()
    if report_type == "custom":
        Submit.place(x=340,y=300)
        Process_btn.place_forget()
        Alarm_btn.place_forget()
        Audit_btn.place_forget()
        pdf_export_btn.place_forget()
    if report_type == "month":
        from_date_time = datetime.today() - timedelta(days=30)
        to_date_time = datetime.today()
        from_date_time1 = from_date_time.strftime("%d-%m-%Y %H:%M:%S")
        to_date_time1 = to_date_time.strftime('%d-%m-%Y %H:%M:%S')
        Label(root,text= "You have selected From date&time  :  "+str(from_date_time1),font = ("Arial", 12)).place (x=30, y= 373)
        Label(root,text= "You have selected To date&time  :  "+str(to_date_time1),font = ("Arial", 12)).place (x=520, y= 373)
        Process_btn.place(x = 50, y = 480)
        Process_btn.config(command = get_process_report)
        pdf_export_btn.place(x=800,y=480)
        pdf_export_btn.config(command = pdf_export)
        Alarm_btn.place (x=300, y = 480)
        Alarm_btn.config(command = generate_alarm_report)
        Audit_btn.place (x=550, y = 480)
        Audit_btn.config(command = generate_audit_report_pdf)
        Submit.place_forget()

## Get PDF Process Report
class get_process_report():
    def __init__ (self):
        global get_all_data_values
        #global btch_filename
        #global path
        global report_type
        global from_date_time,to_date_time,from_date_time1,to_date_time1
        global get_print_interval_time
        print_interval_type = print_interval.get()      
        report_type = report_types_system.get()
        get_print_interval_time = spin_interval.get()
        if print_interval_type == "no_interval":
            if report_type == "custom":
                get_all_data = "select format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total from FloatTable where DateAndTime >= (?) and DateAndTime <= (?) order by DateAndTime asc"
                print("get the point")
                args_strt_end_time = (from_date_time, to_date_time)
                cursor.execute(get_all_data,args_strt_end_time)
                get_all_data_values = cursor.fetchall()
                self.execute_process_report()
            if report_type == "month":
                get_all_data = "select format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total from FloatTable where DateAndTime >= (?) and DateAndTime <= (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time)
                cursor.execute(get_all_data,args_strt_end_time)
                get_all_data_values = cursor.fetchall()
                self.execute_process_report()
        if print_interval_type == "interval":
            if report_type == "custom":
                get_all_data = "SELECT format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total FROM FloatTable where DateAndTime >= (?) and DateAndTime <= (?) AND DATEPART(n, DateAndTime) % 1 = DATEPART(n, DateAndTime) % (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time,get_print_interval_time)
                cursor.execute(get_all_data,args_strt_end_time)
                get_all_data_values = cursor.fetchall()
                self.execute_process_report()
            if report_type == "month":
                get_all_data = "SELECT format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total FROM FloatTable where DateAndTime >= (?) and DateAndTime <= (?) AND DATEPART(n, DateAndTime) % 1 = DATEPART(n, DateAndTime) % (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time,get_print_interval_time)
                cursor.execute(get_all_data,args_strt_end_time)
                get_all_data_values = cursor.fetchall()
                self.execute_process_report() 
    
    def execute_process_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try : 
            if len(get_all_data_values) == 0:
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(get_all_data_values, columns =['Date_time','Value'])
                filename = "Process" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Process_exe" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Process" +".pdf"
                xl_path = "D:\\SCADA\\Format\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                open_new(path1)
        except Exception: 
            showerror ("FileNotFoundError", "File or directory not found")

## Generate Auto Alarm Report
class generate_alarm_report():
    def __init__ (self):
        global from_date_time
        global to_date_time
        global from_date_time1
        global to_date_time1
        global report_type
        self.conn_alarm = pyodbc.connect(
        'DRIVER={SQL Server};'
        'SERVER=IMRANM-LENOVO\WINCC;'
        #'SERVER='+sql_database_name+';'
        'DATABASE=CWI;'
        'UID=SA;'
        'PWD=SA12345;'
        'Trusted_Connection=yes;'
        ) 
        self.cursor_alarm = self.conn_alarm.cursor()

        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        report_type = report_types_system.get()
        if report_type == "custom":
            get_all_data_alarm = "select format(EventTimeStamp,'dd/MM/yyy    HH:mm:ss'),Message,Active from AllEvent where EventTimeStamp >= (?) and EventTimeStamp <= (?) order by EventTimeStamp asc"
            args_strt_end_time = (from_date_time, to_date_time)
            self.cursor_alarm.execute(get_all_data_alarm,args_strt_end_time)
            self.get_all_alarm_value = self.cursor_alarm.fetchall()
            self.execute_alarm_report()
        if report_type == "month":
            get_all_data_alarm = "select format(EventTimeStamp,'dd/MM/yyy    HH:mm:ss'),Message,Active from AllEvent where EventTimeStamp >= (?) and EventTimeStamp <= (?) order by EventTimeStamp asc"
            args_strt_end_time = (from_date_time, to_date_time)
            self.cursor_alarm.execute(get_all_data_alarm,args_strt_end_time)
            self.get_all_alarm_value = self.cursor_alarm.fetchall()
            self.execute_alarm_report()
    
    def execute_alarm_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try :
            if len(self.get_all_alarm_value) == 0:
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(self.get_all_alarm_value, columns =['Date_time','Meassage','Status'])
                filename = "Alarm" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Alarm_exe" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Alarm" +".pdf"
                xl_path = "D:\\SCADA\\Format\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                self.conn_alarm.close()
                open_new(path1)
        except Exception: 
            showerror ("FileNotFoundError", "File or directory not found")

## Generate Auto Audit Report
class generate_audit_report_pdf():
    def __init__ (self):
        global from_date_time,to_date_time,from_date_time1,to_date_time1
        global conn_audit,cursor_audit
        global report_type
        self.conn_audit = pyodbc.connect(
        'DRIVER={SQL Server};'
        'SERVER=IMRANM-LENOVO\WINCC;'
        'DATABASE=Audit;'
        'UID=SA;'
        'PWD=SA12345;'
        'Trusted_Connection=yes;'
        ) 
        self.cursor_audit = self.conn_audit.cursor()
        report_type = report_types_system.get()
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        if report_type == "custom":
            get_all_data_audit = "select format(TimeStmp,'dd/MM/yyy    HH:mm:ss'),MessageText,UserFullName from Auditdata where TimeStmp >= (?) and TimeStmp <= (?) order by TimeStmp asc"
            args_strt_end_time = (from_date_time, to_date_time)
            self.cursor_audit.execute(get_all_data_audit,args_strt_end_time)
            self.get_all_audit_value = self.cursor_audit.fetchall()
            self.execute_audit_report_pdf()
        if report_type == "month":
            get_all_data_audit = "select format(TimeStmp,'dd/MM/yyy    HH:mm:ss'),MessageText,UserFullName from Auditdata where TimeStmp >= (?) and TimeStmp <= (?) order by TimeStmp asc"
            args_strt_end_time = (from_date_time, to_date_time)
            self.cursor_audit.execute(get_all_data_audit,args_strt_end_time)
            self.get_all_audit_value = self.cursor_audit.fetchall()
            self.execute_audit_report_pdf()
    
    def execute_audit_report_pdf(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try:     
            if len(self.get_all_audit_value) == 0:
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(self.get_all_audit_value, columns =['Date_time','Meassage','Username'])
                filename = "Audit" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Audit_exec" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Audit" +".pdf"
                xl_path = "D:\\SCADA\\Format\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                self.cursor_audit.close()
                open_new(path1)
        except Exception: 
            showerror ("FileNotFoundError", "File or directory not found")

## PDF Report Generation
class pdf_export_file():
    def __init__ (self): 
        global from_date_time,to_date_time,from_date_time1,to_date_time1
        global report_types_system_pdf
        global btch_filename
        global pdf_exc_current_dat_time
        pdf_exc_current_dat_time = datetime.today()
        self.pdf_exc_current_dat_time = pdf_exc_current_dat_time.strftime("%d-%m-%Y_%H-%M-%S")
        report_types_system_pdf = report_types_system_pdf.get()
        report_type = report_types_system.get()
        if report_types_system_pdf == "process_pdf" and print_interval_type == "no_interval" :
            if report_type == "custom":
                get_all_data = "select format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total from FloatTable where DateAndTime >= (?) and DateAndTime <= (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time)
                cursor.execute(get_all_data,args_strt_end_time)
                self.get_all_data_values_pdf = cursor.fetchall()
                self.get_pdf_report()
            if report_type == "month":
                get_all_data = "select format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total from FloatTable where DateAndTime >= (?) and DateAndTime <= (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time)
                cursor.execute(get_all_data,args_strt_end_time)
                self.get_all_data_values_pdf = cursor.fetchall()
                self.get_pdf_report()
        if report_types_system_pdf == "process_pdf" and print_interval_type == "interval" :
            if report_type == "custom":               
                get_all_data = "SELECT format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total FROM FloatTable where DateAndTime >= (?) and DateAndTime <= (?) AND DATEPART(n, DateAndTime) % 1 = DATEPART(n, DateAndTime) % (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time,get_print_interval_time)
                cursor.execute(get_all_data,args_strt_end_time)
                self.get_all_data_values_pdf = cursor.fetchall()
                self.get_pdf_report()
            if report_type == "month":
                get_all_data = "SELECT format(DateAndTime,'dd/MM/yyy    HH:mm:ss'),convert(numeric(10,2),Val) as Total FROM FloatTable where DateAndTime >= (?) and DateAndTime <= (?) AND DATEPART(n, DateAndTime) % 1 = DATEPART(n, DateAndTime) % (?) order by DateAndTime asc"
                args_strt_end_time = (from_date_time, to_date_time,get_print_interval_time)
                cursor.execute(get_all_data,args_strt_end_time)
                self.get_all_data_values_pdf = cursor.fetchall()
                self.get_pdf_report()

        if report_types_system_pdf == "alarm_pdf":
            self.conn_alarm = pyodbc.connect(
            'DRIVER={SQL Server};'
            'SERVER=IMRANM-LENOVO\WINCC;'
            'DATABASE=CWI;'
            'UID=SA;'
            'PWD=SA12345;'
            'Trusted_Connection=yes;'
            ) 
            self.cursor_alarm = self.conn_alarm.cursor()

            try:
                os.system('TASKKILL /F /IM AcroRd32.exe')

            except Exception:
                print("KU")
            if report_type == "custom":
                get_all_data_alarm = "select format(EventTimeStamp,'dd/MM/yyy    HH:mm:ss'),Message,Active from AllEvent where EventTimeStamp >= (?) and EventTimeStamp <= (?) order by EventTimeStamp asc"
                args_strt_end_time = (from_date_time, to_date_time)
                self.cursor_alarm.execute(get_all_data_alarm,args_strt_end_time)
                self.get_all_alarm_value_pdf = self.cursor_alarm.fetchall()
                self.get_alarm_pdf_report()
            if report_type == "month":
                get_all_data_alarm = "select format(EventTimeStamp,'dd/MM/yyy    HH:mm:ss'),Message,Active from AllEvent where EventTimeStamp >= (?) and EventTimeStamp <= (?) order by EventTimeStamp asc"
                args_strt_end_time = (from_date_time, to_date_time)
                self.cursor_alarm.execute(get_all_data_alarm,args_strt_end_time)
                self.get_all_alarm_value_pdf = self.cursor_alarm.fetchall()
                self.get_alarm_pdf_report()
        if report_types_system_pdf == "audit_pdf":
            self.conn_audit = pyodbc.connect(
            'DRIVER={SQL Server};'
            'SERVER=IMRANM-LENOVO\WINCC;'
            'DATABASE=Audit;'
            'UID=SA;'
            'PWD=SA12345;'
            'Trusted_Connection=yes;'
            ) 
            self.cursor_audit = self.conn_audit.cursor()
            try:
                os.system('TASKKILL /F /IM AcroRd32.exe')

            except Exception:
                print("KU")
            if report_type == "custom":
                self.get_all_data_audit = "select format(TimeStmp,'dd/MM/yyy    HH:mm:ss'),MessageText,UserFullName from Auditdata where TimeStmp >= (?) and TimeStmp <= (?) order by TimeStmp asc"
                args_strt_end_time = (from_date_time, to_date_time)
                self.cursor_audit.execute(self.get_all_data_audit,args_strt_end_time)
                self.get_all_audit_value = self.cursor_audit.fetchall()
                self.get_audit_pdf_report()
            if report_type == "month":
                self.get_all_data_audit = "select format(TimeStmp,'dd/MM/yyy    HH:mm:ss'),MessageText,UserFullName from Auditdata where TimeStmp >= (?) and TimeStmp <= (?) order by TimeStmp asc"
                args_strt_end_time = (from_date_time, to_date_time)
                self.cursor_audit.execute(self.get_all_data_audit,args_strt_end_time)
                self.get_all_audit_value = self.cursor_audit.fetchall()
                self.get_audit_pdf_report()
    def get_pdf_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try:
            if len(self.get_all_data_values_pdf) == 0:
                pdf_export_scrn.destroy()
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(self.get_all_data_values_pdf, columns =['Date_time','Value'])
                filename = "Process" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Process_exe" +".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Process_Report_" + str(self.pdf_exc_current_dat_time) + ".pdf"
                xl_path = "D:\\PDF\\Process\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                pdf_export_scrn.destroy()
                showinfo ("PDF Export", "PDF Export Successfully")
        except Exception: 
            pdf_export_scrn.destroy()
            showerror ("FileNotFoundError", "File or directory not found")

    def get_alarm_pdf_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try:
            if len(self.get_all_alarm_value_pdf) == 0:
                pdf_export_scrn.destroy()
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(self.get_all_alarm_value_pdf, columns =['Date_time','Meassage','Status'])
                filename = "Alarm" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Alarm_exe" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Alarm_Report_" + str(self.pdf_exc_current_dat_time) +".pdf"
                xl_path = "D:\\PDF\\Alarm\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                self.conn_alarm.close()
                pdf_export_scrn.destroy()
                showinfo ("PDF Export", "PDF Export Successfully")
        except Exception: 
            pdf_export_scrn.destroy()
            showerror ("FileNotFoundError", "File or directory not found")
    def get_audit_pdf_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')
        except Exception:
            print("KU")
        try:    
            if len(self.get_all_audit_value) == 0:
                pdf_export_scrn.destroy()
                showerror ("Error", "No Record Found")
            else:
                df = DataFrame.from_records(self.get_all_audit_value, columns =['Date_time','Meassage','Username'])
                filename = "Audit" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path = xl_path + filename
                wb10 = load_workbook(path,read_only=False)
                ws10 = wb10.worksheets[0]
                c1 = ws10.cell(row=2,column=2)
                c1.value = str(from_date_time1)
                c2 = ws10.cell(row=3,column=2)
                c2.value = str(to_date_time1)
                current_time = datetime.now()
                current_dt_time = current_time.strftime(" %d-%m-%Y  %H:%M:%S")
                c3 = ws10.cell(row=5,column=2)
                c3.value = str(current_dt_time)
                for r in dataframe_to_rows(df,index=False,header=False):
                    ws10.append(r)
                filename = "Audit_exec" + ".xlsx"
                xl_path = "D:\\SCADA\\Format\\"
                path3 = xl_path + filename
                wb10.save(path3)
                sleep(1)
                excel = client.DispatchEx("Excel.Application")
                excel.Visible = 0
                wb = excel.Workbooks.Open(path3)
                ws = wb.Worksheets[0]
                filename = "Audit_Report_" + str(self.pdf_exc_current_dat_time) +".pdf"
                xl_path = "D:\\PDF\\Audit\\"
                path1 = xl_path + filename
                ws.ExportAsFixedFormat(0,path1)
                wb.Close()
                excel.Quit()
                self.cursor_audit.close()
                pdf_export_scrn.destroy()
                showinfo ("PDF Export", "PDF Export Successfully")
        except Exception: 
            pdf_export_scrn.destroy()
            showerror ("FileNotFoundError", "File or directory not found")

## PDF Export Popup Screen
class pdf_export():
    def __init__ (self):
        global report_types_system_pdf,get_print_interval_time,pdf_export_scrn,print_interval_type,report_type,report_types_system_pdf       
        HighlightFont_pdf = font.Font(family='Helvetica', size=12, weight='bold')
        pdf_export_scrn = Toplevel(root)
        pdf_export_scrn.wm_attributes("-transparentcolor", 'grey')
        pdf_export_scrn.title("PDF EXPORT")
        x_position = 710
        y_position = 200
        pdf_export_scrn.geometry(f"250x250+{x_position}+{y_position}")
        report_types_system_pdf = StringVar(pdf_export_scrn,"process_pdf")
        print_interval_type = print_interval.get()      
        report_type = report_types_system.get()
        get_print_interval_time = spin_interval.get()
        def cancel_pdf_selection():
            pdf_export_scrn.destroy()
        report_type_pdf = Radiobutton(pdf_export_scrn, text = "PROCESS",font= HighlightFont_pdf,height=2, width=15,indicatoron=1,variable = report_types_system_pdf, value = "process_pdf")
        report_type_pdf.place (x=10,y = 50)
        report_type_pdf = Radiobutton(pdf_export_scrn, text = "ALARM", font= HighlightFont_pdf,height=2, width=15,indicatoron=1,variable = report_types_system_pdf, value ="alarm_pdf")
        report_type_pdf.place (x=0,y = 100)
        report_type_pdf = Radiobutton(pdf_export_scrn, text = "AUDIT",font= HighlightFont_pdf,height=2, width=15,indicatoron=1, variable = report_types_system_pdf, value ="audit_pdf")
        report_type_pdf.place (x=0,y = 150)
        Label (pdf_export_scrn,text = "Select the Report to Export PDF",font = ("Arial", 12)).place (x=10, y = 10)
        btn_to_export_pdf = Button(pdf_export_scrn,text = "Save",font= HighlightFont_pdf,height=1, width=8,foreground="white",background="blue",highlightthickness=2,highlightcolor="white",command=pdf_export_file)
        btn_to_export_pdf.place(x=20,y=200)
        btn_to_export_pdf_cncl = Button(pdf_export_scrn,text = "Cancel",font= HighlightFont_pdf,height=1, width=8,foreground="white",background="red",highlightthickness=2,highlightcolor="white",command=cancel_pdf_selection)
        btn_to_export_pdf_cncl.place(x=140,y=200)


## Date Time Selection Customize
def date_selc():
    global custom_to_dttime
    global custom_from_dttime
    ## get from time and date 
    hr_get = hour_spin.get()
    min_get = min_spin.get()
    sec_get = sec_spin.get()
    hours = 25
    minutes = 61
    seconds = 61
    global from_value
    try :
        hr_get = int(hr_get) 
    
        min_get = int(min_get)
        sec_get = int(sec_get)
        if (hours > hr_get) and (minutes > min_get) and (seconds > sec_get):
            print("success")
        else :
            showerror ("Error", "Invalid Entry")

    except Exception:
        showerror ("Error", "Invalid Entry")
    else :
        str_hr = hour_spin.get()
        str_min = min_spin.get()
        str_sec =  sec_spin.get()
        date_value = selec_date.get_date()
        date_selc = date_value.strftime("%d-%m-%Y")
        from_value = str(date_selc) + " " + str(str_hr) + ":" + str(str_min) + ":" + str(str_sec)
        custom_from_dttime = Label(root, text = "You have selected From date&time  :  " + str(date_selc) + " "+str(str_hr) + ":" +str(str_min)+ ":"+str(str_sec),font = ("Arial", 12))
        custom_from_dttime.place (x=30, y= 373)
    ## get to date and time
    to_hr_get = to_hour_spin.get()
    to_min_get = to_min_spin.get()
    to_sec_get = to_sec_spin.get()
    hours = 25
    minutes = 61
    seconds = 61
    try :
        to_hr_get = int(to_hr_get) 
    
        to_min_get = int(to_min_get)
        to_sec_get = int(to_sec_get)
        if (hours > to_hr_get) and (minutes > to_min_get) and (seconds > to_sec_get):
            print("success")
        else :
            showerror ("Error", "Invalid Entry")

    except Exception:
        showerror ("Error", "Invalid Entry")
    else :
        to_str_hr = to_hour_spin.get()
        to_str_min = to_min_spin.get()
        to_str_sec =  to_sec_spin.get()
        to_date_value = to_selec_date.get_date()
        #print(to_date_value)
        to_date_selc = to_date_value.strftime("%d-%m-%Y")
        to_value = str(to_date_selc) + " " + str(to_str_hr) + ":" + str(to_str_min) + ":" + str(to_str_sec)
        custom_to_dttime = Label(root, text = "You have selected To date&time  :  " +str(to_date_selc) + " "+str(to_str_hr) + ":" +str(to_str_min)+ ":"+str(to_str_sec),font = ("Arial", 12))
        custom_to_dttime.place (x=520, y= 373)
        #print(to_value)
    global from_date_time,to_date_time,from_date_time1,to_date_time1
    from_date_time = datetime.strptime(from_value,'%d-%m-%Y %H:%M:%S')
    to_date_time = datetime.strptime(to_value,'%d-%m-%Y %H:%M:%S')
    from_date_time1 = from_date_time.strftime("%d-%m-%Y %H:%M:%S:%S")
    to_date_time1 = to_date_time.strftime('%d-%m-%Y %H:%M:%S:%S')
    try :
        if (from_date_time >= to_date_time):
            #print("error")
            Process_btn.place_forget()
            Alarm_btn.place_forget()
            Audit_btn.place_forget()
            pdf_export_btn.place_forget()
            raise showerror("Error", "TO DATE SHOULD BE GREATER THEN FROM DATE")                    
    except :
        print("error")
    else:
        Process_btn.place(x = 50, y = 480)
        Process_btn.config(command = get_process_report)
        pdf_export_btn.place(x=800,y=480)
        pdf_export_btn.config(command = pdf_export)
        Alarm_btn.place (x=300, y = 480)
        Alarm_btn.config(command = generate_alarm_report)
        Audit_btn.place (x=550, y = 480)
        Audit_btn.config(command = generate_audit_report_pdf)


def main1():
    global selec_date,from_min_selc,hour_spin,min_spin,sec_spin
    global to_selec_date,to_hour_spin,to_min_spin,to_sec_spin
    global Process_btn,Alarm_btn,Audit_btn,report_types_system
    global report_type
    global Submit
    global print_interval,spin_interval,interval_label,get_print_interval_time
    global pdf_export_btn
    interval_selc = IntVar(root)
    report_types_system = StringVar(root,"custom")
    print_interval = StringVar(root,"no_interval")
    HighlightFont = font.Font(family='Helvetica', size=12, weight='bold')
    from_hr_selc = IntVar(root)
    from_min_selc = IntVar(root)
    from_sec_selc = IntVar(root) 
    to_hr_selc = IntVar(root)
    to_min_selc = IntVar(root)
    to_sec_selc = IntVar(root)
    ## From Date Time Selection
    Label(root,text = "Choose From Date and Time" ,font = ("Arial", 15)).place(x = 30, y = 160)
    selec_date = DateEntry(root, font = "Arial 14", selectmode = 'day')
    selec_date.place(x = 30, y = 200)
    Label (root,text = "Hours",font = ("Arial", 15)).place (x=30, y = 320)
    hour_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24"),textvariable = from_hr_selc,width=2,font=('verdana',30))
    hour_spin.place(x = 30, y = 250)
    Label (root,text = "Minutes",font = ("Arial", 15)).place (x=130, y = 320)
    min_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = from_min_selc,width=2,font=('verdana',30))
    min_spin.place(x = 130, y = 250)
    Label (root,text = "Seconds",font = ("Arial", 15)).place (x=220, y = 320)
    sec_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = from_sec_selc,width=2,font=('verdana',30))
    sec_spin.place(x = 220, y = 250)
    ## to date and time selection
    Label(root,text = "Choose To Date and Time" ,font = ("Arial", 15)).place(x = 520, y = 160)
    to_selec_date = DateEntry(root, font = "Arial 14", selectmode = 'day')
    to_selec_date.place(x = 520, y = 200)
    Label (root,text = "Hours",font = ("Arial", 15)).place (x=520, y = 320)
    to_hour_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24"),textvariable = to_hr_selc,width=2,font=('verdana',30))
    to_hour_spin.place(x = 520, y = 250)

    Label (root,text = "Minutes",font = ("Arial", 15)).place (x=610, y = 320)
    to_min_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = to_min_selc,width=2,font=('verdana',30))
    to_min_spin.place(x = 610, y = 250)
    Label (root,text = "Seconds",font = ("Arial", 15)).place (x=710, y = 320)
    to_sec_spin = Spinbox(root, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = to_sec_selc,width=2,font=('verdana',30))
    to_sec_spin.place(x = 710, y = 250)
    Submit = Button(root,text = 'SUBMIT',font= HighlightFont,height=2, width=13,command=date_selc)
    Submit.place(x=340,y=300)
    Label(root,text = "                                                                 DATE TIME PICKER                                                                                                ",bg='navy',fg='white',bd=8,font = ("Arial", 15,font.BOLD)).place(x=0,y=110)
    Label(root,text = "                                                                 REPORT ANALYSIS                                                                                                 ",bg='navy',fg='white',bd=8,font = ("Arial", 15,font.BOLD)).place(x=0,y=420)
    Label(root,text = "                                                       SELECT CUSTOM REPORT OR DETAIL REPORT                                                                             ",bg='navy',fg='white',bd=8,font = ("Arial", 15,font.BOLD)).place(x=0,y=1)   
    report_type = Radiobutton(root, text = "Custom Date",font = ("Arial", 15),height=2, width=15,indicatoron=1,variable = report_types_system, value = "custom",command=selection_of_report)
    report_type.place (x=150, y = 50)
    report_type = Radiobutton(root, text = "Last Month", font = ("Arial", 15),height=2, width=15,indicatoron=1,variable = report_types_system, value ="month",command=selection_of_report)
    report_type.place (x=450, y = 50)  
    Process_btn = Button(root, text = "PROCESS REPORT",font= HighlightFont,height=2, width=15)
    Alarm_btn = Button(root, text = "ALARM REPORT",font= HighlightFont,height=2, width=15)
    Audit_btn = Button(root, text = "AUDIT REPORT",font= HighlightFont,height=2, width=15)    
    Label(root,text = "Select Print Interval" ,font = ("Arial", 15)).place(x = 850, y = 160)    
    report_interval = Radiobutton(root, text = "Minute Interval",font= HighlightFont,height=2, width=15,indicatoron=1, variable = print_interval, value ="interval",command=print_interval_selc)
    report_interval.place (x=800, y = 190)
    report_interval = Radiobutton(root, text = "No Interval",font= HighlightFont,height=2, width=15,indicatoron=1, variable = print_interval, value ="no_interval",command=print_interval_selc)
    report_interval.place (x=960, y = 190)
    interval_label = Label(root,text = "Minute" ,font = ("Arial", 15))
    spin_interval = Spinbox(root, values = ("01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = interval_selc,width=2,font=('verdana',30))
    pdf_export_btn = Button(root, text = "PDF EXPORT",font= HighlightFont,height=2, width=15)
    root.mainloop()

if(__name__=="__main__"): 
    main1()
