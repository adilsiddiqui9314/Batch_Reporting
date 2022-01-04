from tkinter import*
import tkinter as tk 
from tkinter import ttk
import os, glob, time
import webbrowser as wb
from tkinter.messagebox import showerror
from tkinter import font
from tkcalendar import DateEntry
import pyodbc
from openpyxl import load_workbook,Workbook
from time import sleep
from win32com import client
import threading
from tkinter import messagebox as mb
import datetime as dt
from datetime import datetime


root = Tk()
root.title("REPORT VIEWER")
root.state('zoomed')
wrapper1 = LabelFrame(root, text = "SELECT DATA ")
wrapper1.place(x=10, y =10)
scrol_y = Scrollbar(wrapper1,orient = VERTICAL)
scrol_x = Scrollbar(wrapper1,orient = HORIZONTAL)

my_tree = ttk.Treeview(wrapper1,xscrollcommand=scrol_x.set,yscrollcommand= scrol_y.set, columns=[1,2,3],show = "headings",height = "20",)

my_tree.heading(1,text = "BATCHNUMBER")
my_tree.heading(2,text = "LOTNUMBER")
my_tree.heading(3,text = "BATCH DATETIME")
my_scrollbar = Scrollbar(wrapper1,orient = VERTICAL)
scrol_y.config(command=my_tree.yview)
scrol_y.pack(side=RIGHT,fill=Y)
scrol_x.config(command=my_tree.xview)
scrol_x.pack(side=BOTTOM,fill=X)

my_tree.pack()
#global my_tree,sytem_selected,path,btch_filename
global batchname_rmg,lotnumber_rmg
global batchname_fbd,lotnumber_fbd
global batchname_bld,lotnumber_bld
try:
    os.system('TASKKILL /F /IM excel.exe')

except Exception:
    print("KU")
conn = pyodbc.connect(
    'DRIVER={SQL Server};'
    'SERVER=GLENMARK\SQLEXPRESS;'
    'DATABASE=GT;'
    'Trusted_Connection=yes;'
) 

cursor = conn.cursor()

class rmg_report_fetch():
    def __init__ (self):
        global sytem_selected
        generate_report_fbd_auto.place_forget(),generate_report_fbd_alarm.place_forget(),generate_report_fbd_audit.place_forget()
        ent_search_fbd.place_forget(),clear_fbd_record.place_forget(),search_fbd_data.place_forget()
        generate_report_bld_auto.place_forget(),generate_report_bld_alarm.place_forget(),generate_report_bld_audit.place_forget()
        ent_search_bld.place_forget(),clear_bld_record.place_forget(),search_bld_data.place_forget()
        alrm_btn.place_forget(),alrm_aud_btn.place_forget()       
        sytem_selected = report_types_system.get()
        my_tree.delete(*my_tree.get_children())
        ## get distinct rmg batchname and lotnumber
        commands = ("select distinct OpBatchName,OpLotNumber,Max(date_time) as BATCHLOT from RMGReportAuto group by OpBatchName,OpLotNumber order by Max(date_time) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()

        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))

        def val (event):
            global select_report   
            global batchname_rmg
            global lotnumber_rmg
            #item = my_tree.item(my_tree.focus())
            item1 = my_tree.focus()
            item = my_tree.item(item1,option="values")
            #batchname_rmg1 = (item['values'][0])
            batchname_rmg1 = item[0]
            batchname_rmg = str(batchname_rmg1)
            lotnumber_rmg1 = item[1]
            lotnumber_rmg = str(lotnumber_rmg1)
            generate_report_rmg_auto.place(x= 650,y=70)
            generate_report_rmg_alarm.place(x = 850 , y = 70)
            generate_report_rmg_audit.place(x = 1050, y = 70)
            generate_report_rmg_auto.config(command = self.rmg_auto_rep_execute)
            generate_report_rmg_alarm.config(command = self.rmg_auto_alarm_execute)
            generate_report_rmg_audit.config(command = self.rmg_auto_audit_execute)
        
        #my_tree.bind('<Double-1>',val)
        my_tree.bind('<ButtonRelease-1>',val)
        search_rmg_data.place(x= 1000,y=200)
        lbl_search.place(x = 650, y = 200)
        ent_search.place(x = 750, y = 210)
        keyboard_btn.place(x = 650, y = 170)
        clear_rmg_record.place( x= 1150, y= 200)
        alrm_btn.place(x=750, y = 250)
        alrm_aud_btn.place(x=1050, y = 250)
        alrm_btn.config(command = report_fetch)
        alrm_aud_btn.config(command = report_fetch)
        show_aud_alrm_reprt.config(command = audit_alarm_reports)
        search_rmg_data.config(command = self.search_rmg_data)
        clear_rmg_record.config(command = self.clear_records)
    ## search RMG distinct value
    def search_rmg_data(self):
        get_search_rmg_value = ent_search.get()
        querry = "select distinct OpBatchName,OpLotNumber as BATCHLOT from RMGReportAuto where (OpBatchName = (?)) OR (OpLotNumber = (?)) "
        cursor.execute(querry,get_search_rmg_value,get_search_rmg_value)
        get_filter_vlve=cursor.fetchall()
        conn.commit()
        if len(get_filter_vlve)==0:
            my_tree.delete(*my_tree.get_children())
            my_tree.insert('',index='end',values = ("NO RECORD FOUND"))
        else:
            my_tree.delete(*my_tree.get_children())
            for values in get_filter_vlve:
                batchname = str(values[0])
                lotnumber = str(values[1])
                my_tree.insert('',index='end',values = (batchname,lotnumber))
    
    ## clear search rmg records
    def clear_records(self):
        my_tree.delete(*my_tree.get_children())
        commands = ("select distinct OpBatchName,OpLotNumber, Max(date_time) as BATCHLOT from RMGReportAuto group by OpBatchName,OpLotNumber order by max(date_time) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()
        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))    
    def rmg_auto_rep_execute(self):
        confirm_before_export_rmg_report=mb.askquestion('Export File', 'Are you Sure you want to View the RMG Batch Report')
        if confirm_before_export_rmg_report == "yes" :
            threading.Thread(target= self.generate_rmg_report).start()
    def rmg_auto_alarm_execute(self):
        confirm_before_export_rmg_alarm=mb.askquestion('Export File', 'Are you Sure you want to View the RMG ALARM Report')
        if confirm_before_export_rmg_alarm == "yes" :
            threading.Thread(target= self.generate_rmg_alarm).start()
    def rmg_auto_audit_execute(self):
        confirm_before_export_rmg_audit=mb.askquestion('Export File', 'Are you Sure you want to View the RMG AUDIT Report')
        if confirm_before_export_rmg_audit == "yes" :
            threading.Thread(target= self.generate_rmg_audit).start()
    
    def generate_rmg_report(self):
        #commands1 = ("SELECT format(date_time,'dd/mm/yyy    hh:mm:ss'),OpAgiRpmAct,OpAgiSlowRPM,OpChopRpmAct,OpChopSlowRPM,OpAgiCurr,OpChopCurr,OpDosingRPM,OpDosingRPMSV,OpComillRpmAct,OpWetComillRPM,OpProdTempAct,OpRemark from RMGReportAuto where (OpBatchName = (?)) AND (OpLotNumber = (?)) AND (OpRemark = 'RMG PROCESS STARTED') order by date_time") 
        #cursor.execute(commands1,batchname_rmg,lotnumber_rmg)
        #rows1=cursor.fetchall()
        #conn.commit()
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        commands2 = ("SELECT format(date_time,'dd/MM/yyy    HH:mm:ss'),OpAgiRPMSV,OpAgiRpmAct,OpChopRPMSV,OpChopRpmAct,convert(numeric(10,1),OpAgiCurr) as Total,convert(numeric(10,1),OpChopCurr) as Total,OpDosingRPMSV,OpDosingRPM,OpWetComillRPM,OpComillRpmAct,convert(numeric(10,1),OpProdTempAct) as Total,OpRemark from RMGReportAuto where (OpBatchName = (?)) AND (OpLotNumber = (?)) AND (OpRemark = 'RMG PROCESS STARTED' or OpRemark = 'RMG PROCESS RUNNING' or OpRemark = 'RMG ADD BINDER' or OpRemark = 'WET MIXING CYCLE 2 STARTED' or OpRemark = 'WET MIXING CYCLE 3 STARTED' or OpRemark = 'DISCHARGE CYCLE STARTED' or OpRemark = 'RMG DISCHARGE PROCESS RUNNING' or OpRemark = 'RMG PROCESS COMPLETED') order by date_time") 
        cursor.execute(commands2,batchname_rmg,lotnumber_rmg)
        rows2=cursor.fetchall()
        conn.commit()
        batch_data =  rows2

        commands3 = ("SELECT MIN(format(date_time,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time,'dd/MM/yyy    HH:mm:ss')) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands3,batchname_rmg,lotnumber_rmg)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

                    ##             0       1           2           3       4           5           6           7               8                       9                      10              11                  12                  13                    14                15                 16               17                      18                        19                      20                  21             22                   23                      24              25                  26                  27                      28                          29                      30                  31              32                  33                      34               35                36                  37                      38                      39                    40                  41           42          43              44          45          46                                          
        commands4=("Select OpProdName,OpProdCode,OpBatchName,OpLotNumber,OpEqpID,OpBatchSize,OpPrintInterval,OpAgiSlowDryTime,OpAgiSlowDryTime_sec,OpAgiFastDryTime,OpAgiFastDryTime_sec,Op_Agi_Wet_Slow_Time,Op_Agi_Wet_Fast_Time,Op_Chop_Slow_Time,Op_Chop_Fast_Time,OpChopDelayTime,Op_Agi_Wet_Slow_Time_Sec,Op_Agi_Wet_Fast_Time_Sec,Op_Chop_Slow_Time_Sec,Op_Chop_Fast_Time_Sec,OpDosingTime,Op_Agi_Wet_Slow_Time2,Op_Agi_Wet_Fast_Time2,Op_Chop_Slow_Time2,Op_Chop_Fast_Time2,OpChopDelayTime2,Op_Agi_Wet_Slow_Time2_Sec,Op_Agi_Wet_Fast_Time2_Sec,Op_Chop_Slow_Time2_Sec,Op_Chop_Fast_Time2_Sec,OpDosingTime2,Op_Agi_Wet_Slow_Time3,Op_Agi_Wet_Fast_Time3,Op_Chop_Slow_Time3,Op_Chop_Fast_Time3,OpChopDelayTime3,Op_Agi_Wet_Slow_Time3_Sec,Op_Agi_Wet_Fast_Time3_Sec,Op_Chop_Slow_Time3_Sec,Op_Chop_Fast_Time3_Sec,OpDosingTime3,OpAgiSlowRPM,OpAgiFastRPM,OpChopSlowRPM,OpChopFastRPM,OpDosingRPMSV,OpWetComillRPM from RMGReportAuto where (OpBatchName = (?)) and (OpLotNumber = (?)) and (OpRemark = 'RMG PROCESS STARTED' OR OpRemark = 'RMG PROCESS RUNNING') order by date_time")
        cursor.execute(commands4,batchname_rmg,lotnumber_rmg)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands_user_time = ("Select MIN(date_time) AS StartTime,MAX(date_time) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands_user_time,batchname_rmg,lotnumber_rmg)
        rows_user=cursor.fetchall()
        conn.commit()
        sort_strt_end_time_rmg = [item for t in rows_user for item in t]
        start_time_user_rmg = sort_strt_end_time_rmg[0]
        end_time_user_rmg = sort_strt_end_time_rmg[1]

        commands5 = ("Select OpUsername,date1,time1 from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands5,batchname_rmg,lotnumber_rmg,start_time_user_rmg)
        batch_start_user=cursor.fetchone()
        conn.commit()
        commands6 = ("Select OpUsername,date1,time1 from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands6,batchname_rmg,lotnumber_rmg,end_time_user_rmg)
        batch_end_user=cursor.fetchone()
        conn.commit()
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        wb = load_workbook('E:\\SCADA\\Format\\RMG_Auto_Batch_Report.xlsm',read_only=False,keep_vba=True)
        ws = wb.worksheets[0]
        a = 3
        b = 17
        c= 17
        d=23
        e=23
        f=29
        g=29
        h=42
        OpAgiSlowDryTime = rows_sp[7]
        OpAgiSlowDryTime_sec = rows_sp[8]
        OpAgiFastDryTime = rows_sp[9]
        OpAgiFastDryTime_sec = rows_sp[10]
        OpDosingTime = rows_sp[20]
        OpDosingTime2 = rows_sp[30]
        OpDosingTime3 = rows_sp[40]
        OpAgiSlowRPM = rows_sp[41]
        OpAgiFastRPM = rows_sp[42]
        OpChopSlowRPM = rows_sp[43]
        OpChopFastRPM = rows_sp[44]
        OpDosingRPMSV = rows_sp[45]
        OpWetComillRPM = rows_sp[46]
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        ## recipe name loop
        for x in rows_sp[0:7] :
            c1 = ws.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1

        c2 = ws.cell(row=15,column=2)
        c2.value = str(OpAgiSlowDryTime)
        c3 = ws.cell(row=15,column=4)
        c3.value = str(OpAgiSlowDryTime_sec)
        c4 = ws.cell(row=16,column=2)
        c4.value = str(OpAgiFastDryTime)
        c5 = ws.cell(row=16,column=4)
        c5.value = str(OpAgiFastDryTime_sec)
        c14 = ws.cell(row=11,column=2)
        c14.value = str(start_time)
        c15 = ws.cell(row=9,column=7)
        c15.value = str(end_time)
        c16 = ws.cell(row=50,column=2)
        c16.value = str(start_time)
        c17 = ws.cell(row=48,column=7)
        c17.value = str(end_time)
        c18 = ws.cell(row=12,column=2)
        c18.value = str(batch_start_user)
        c19 = ws.cell(row=10,column=7)
        c19.value = str(batch_end_user)
        c20 = ws.cell(row=51,column=2)
        c20.value = str(batch_start_user)
        c21 = ws.cell(row=49,column=7)
        c21.value = str(batch_end_user)
        c22 = ws.cell(row=11,column=7)
        c22.value = str(current_dt_time)
        c23 = ws.cell(row=50,column=7)
        c23.value = str(current_dt_time)
        current_user_cell = ws.cell(row=12,column=7)
        current_user_cell.value = str(current_user)
        current_user_cell2 = ws.cell(row=51,column=7)
        current_user_cell2.value = str(current_user)
        ## wet mixing 1 minute loop
        for x in rows_sp[11:16] :
            c6 = ws.cell(row=b+1,column=2)
            c6.value = str(x)
            b+=1
        ## wet mixing 1 seconds loop
        for x in rows_sp[16:20] :
            c7 = ws.cell(row=c+1,column=4)
            c7.value = str(x)
            c+=1

        c8 = ws.cell(row=22,column=7)
        c8.value = str(OpDosingTime)
        ## wet mixing 2 minute loop
        for x in rows_sp[21:26] :
            c9 = ws.cell(row=d+1,column=2)
            c9.value = str(x)
            d+=1
        ## wet mixing 2 seconds loop
        for x in rows_sp[26:30] :
            c9 = ws.cell(row=e+1,column=4)
            c9.value = str(x)
            e+=1

        c10 = ws.cell(row=28,column=7)
        c10.value = str(OpDosingTime2)
        ## wet mixing 3 minute loop
        for x in rows_sp[31:36] :
            c10 = ws.cell(row=f+1,column=2)
            c10.value = str(x)
            f+=1
        ## wet mixing 3 seconds loop
        for x in rows_sp[36:40] :
            c11 = ws.cell(row=g+1,column=4)
            c11.value = str(x)
            g+=1

        c12 = ws.cell(row=34,column=7)
        c12.value = str(OpDosingTime3)
        ## RPM insert row and column
        c45 = ws.cell(row=36,column=2)
        c45.value = str(OpAgiSlowRPM)
 
        c46 = ws.cell(row=37,column=2)
        c46.value = str(OpAgiFastRPM)

        c49 = ws.cell(row=38,column=2)
        c49.value = str(OpDosingRPMSV)

        c47 = ws.cell(row=36,column=7)
        c47.value = str(OpChopSlowRPM)

        c48 = ws.cell(row=37,column=7)
        c48.value = str(OpChopFastRPM)

        c50 = ws.cell(row=38,column=7)
        c50.value = str(OpWetComillRPM)

        ## recipe name loop of second page
        for x in rows_sp[0:7] :
            c1 = ws.cell(row=h+1,column=2)
            c1.value = str(x)
            h+=1
        ## process interval insert data
        p = 54
        for i in batch_data:
            k = 0
            for j in i:
                c25 = ws.cell(row=p+1,column=k+1)
                c25.value = str(j)
                k+=1
            p+=1
        sleep(1)
        btch_filename_rmg_alarm = "trail_testing1"
        filename = "RMG_Auto_Batch_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\RMG\\EXEC\\"
        xl_path1 = xl_path + filename
        wb.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\RMG\\EXEC\\RMG_Auto_Batch_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb.close()
        sleep(2)
        root.destroy()

    def generate_rmg_alarm(self):

        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")

        commands2= ("SELECT MIN(date_time) AS StartTime, MAX(date_time) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands2,batchname_rmg,lotnumber_rmg)
        rmg_alarm_start_end_time=cursor.fetchall()
        sort_strt_end_time = [item for t in rmg_alarm_start_end_time for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(date_time,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time,'dd/MM/yyy    HH:mm:ss')) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands3,batchname_rmg,lotnumber_rmg)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]
                    ##           0           1          2           3       4          5           6                7         
        commands4=("Select OpProdName,OpProdCode,OpBatchName,OpLotNumber,OpEqpID,OpBatchSize,OpPrintInterval,OpAgiSlowDryTime from RMGReportAuto where (OpBatchName = (?)) and (OpLotNumber = (?)) and (OpRemark = 'RMG PROCESS STARTED' OR OpRemark = 'RMG PROCESS RUNNING') order by date_time")
        cursor.execute(commands4,batchname_rmg,lotnumber_rmg)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands5 = ("Select OpUsername from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands5,batchname_rmg,lotnumber_rmg,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select OpUsername from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands6,batchname_rmg,lotnumber_rmg,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user))

        alarm_value_fetch = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) and Al_Group = 1 order by Al_Start_Time"
        cursor.execute(alarm_value_fetch,start_time,end_time)
        alarm_values=cursor.fetchall()
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        wb1 = load_workbook('E:\\SCADA\\Format\\RMG_Auto_Alarm_Report.xlsm',read_only=False,keep_vba=True)
        ws1 = wb1.worksheets[0]
        btch_filename_rmg_alarm = "testing_alarm"
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws1.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws1.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws1.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws1.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws1.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws1.cell(row=12,column=4)
        c24.value = str(current_dt_time)        
        current_user_cell = ws1.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(alarm_values)==0:
            c17 = ws1.cell(row=14,column=3)
            c17.value = str("NO ALARM FOUND REPORT")
        else:
            for i in alarm_values:
                k = 0
                for j in i:
                    c25 = ws1.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "RMG_Auto_Alarm_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\RMG\\EXEC\\"
        xl_path1 = xl_path + filename
        wb1.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\RMG\\EXEC\\RMG_Auto_Alarm_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb1.close()
        sleep(2)
        root.destroy()

    def generate_rmg_audit(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")

        commands3_audit = ("SELECT MIN(date_time) AS StartTime, MAX(date_time) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands3_audit,batchname_rmg,lotnumber_rmg)
        commands3_audit=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in commands3_audit for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(date_time,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time,'dd/MM/yyy    HH:mm:ss')) FROM RMGReportAuto WHERE OpBatchName = (?) AND OpLotNumber = (?)")
        cursor.execute(commands3,batchname_rmg,lotnumber_rmg)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]
                            ##           0           1          2           3       4          5           6                7         
        commands4=("Select OpProdName,OpProdCode,OpBatchName,OpLotNumber,OpEqpID,OpBatchSize,OpPrintInterval,OpAgiSlowDryTime from RMGReportAuto where (OpBatchName = (?)) and (OpLotNumber = (?)) and (OpRemark = 'RMG PROCESS STARTED' OR OpRemark = 'RMG PROCESS RUNNING') order by date_time")
        cursor.execute(commands4,batchname_rmg,lotnumber_rmg)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands5 = ("Select OpUsername from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands5,batchname_rmg,lotnumber_rmg,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select OpUsername from RMGReportAuto where OpBatchName = (?) and OpLotNumber = (?) and date_time = (?)")
        cursor.execute(commands6,batchname_rmg,lotnumber_rmg,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        audit_value_fetch = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'FBD%') AND (Ev_Message NOT LIKE 'BLENDER%') order by Ev_Time"
        cursor.execute(audit_value_fetch,start_time,end_time)
        audit_value_fetch=cursor.fetchall()
        btch_filename_rmg_alarm = "Audit_testing1"
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        wb2 = load_workbook('E:\\SCADA\\Format\\RMG_Auto_Audit_Report.xlsm',read_only=False,keep_vba=True)
        ws2 = wb2.worksheets[0]
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws2.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws2.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws2.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws2.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws2.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws2.cell(row=12,column=4)
        c24.value = str(current_dt_time)
        current_user_cell = ws2.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(audit_value_fetch)==0:
            c17 = ws2.cell(row=14,column=2)
            c17.value = str("NO AUDIT FOUND REPORT")
        else:
            for i in audit_value_fetch:
                k = 0
                for j in i:
                    c25 = ws2.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "RMG_Auto_Audit_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\RMG\\EXEC\\"
        xl_path1 = xl_path + filename
        wb2.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\RMG\\EXEC\\RMG_Auto_Audit_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb2.close()
        sleep(2)
        root.destroy()

class fbd_report_fetch():
    def __init__ (self):
        global sytem_selected
        sytem_selected = report_types_system.get()
        generate_report_rmg_auto.place_forget(),generate_report_rmg_alarm.place_forget(),generate_report_rmg_audit.place_forget()
        ent_search.place_forget(),clear_rmg_record.place_forget(),search_rmg_data.place_forget()
        generate_report_bld_auto.place_forget(),generate_report_bld_alarm.place_forget(),generate_report_bld_audit.place_forget()
        ent_search_bld.place_forget(),clear_bld_record.place_forget(),search_bld_data.place_forget()
        alrm_aud_btn.place_forget()
        alrm_btn.place_forget()
        ent_search.place_forget()
        clear_rmg_record.place_forget()
        search_rmg_data.place_forget()
        my_tree.delete(*my_tree.get_children())
        commands = ("select distinct FbdBatchName,FbdLotNumber, Max(date_time2) as BATCHLOT from FbpReportAuto group by FbdBatchName,FbdLotNumber order by max(date_time2) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()

        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))
        def val (event):
            global select_report   
            global batchname_fbd
            global lotnumber_fbd
            item1 = my_tree.focus()
            item = my_tree.item(item1,option="values")
            batchname_fbd1 = item[0]
            batchname_fbd = str(batchname_fbd1)
            lotnumber_fbd1 = item[1]
            lotnumber_fbd = str(lotnumber_fbd1)
            generate_report_fbd_auto.place(x= 650,y=70)
            generate_report_fbd_alarm.place(x = 850 , y = 70)
            generate_report_fbd_audit.place(x = 1050, y = 70)

            generate_report_fbd_auto.config(command = self.fbd_auto_rep_execute)
            generate_report_fbd_alarm.config(command = self.fbd_auto_alarm_execute)
            generate_report_fbd_audit.config(command = self.fbd_auto_audit_execute)
        #my_tree.bind('<Double-1>',val)
        my_tree.bind('<ButtonRelease-1>',val)
        search_fbd_data.place(x= 1000,y=200)
        lbl_search.place(x = 650, y = 200)
        ent_search_fbd.place(x = 750, y = 210)
        keyboard_btn.place(x = 650, y = 170)
        clear_fbd_record.place( x= 1150, y= 200)
        alrm_btn.place(x=750, y = 250)
        alrm_aud_btn.place(x=1050, y = 250)
        alrm_btn.config(command = report_fetch)
        alrm_aud_btn.config(command = report_fetch)
        show_aud_alrm_reprt.config(command = audit_alarm_reports)
        search_fbd_data.config(command = self.search_fbd_data)
        clear_fbd_record.config(command = self.clear_records_fbd)
    ## search FBD distinct value
    def search_fbd_data(self):
        get_search_rmg_value = ent_search_fbd.get()
        querry = "select distinct FbdBatchName,FbdLotNumber as BATCHLOT from FbpReportAuto where (FbdBatchName = (?)) OR (FbdLotNumber = (?)) "
        cursor.execute(querry,get_search_rmg_value,get_search_rmg_value)
        get_filter_vlve=cursor.fetchall()
        conn.commit()
        #print(get_filter_vlve)
        if len(get_filter_vlve)==0:
            my_tree.delete(*my_tree.get_children())
            my_tree.insert('',index='end',values = ("NO RECORD FOUND"))
        else:
            my_tree.delete(*my_tree.get_children())
            for values in get_filter_vlve:
                batchname = str(values[0])
                lotnumber = str(values[1])
                my_tree.insert('',index='end',values = (batchname,lotnumber))
    
    ## clear search rmg records
    def clear_records_fbd(self):
        my_tree.delete(*my_tree.get_children())
        commands = ("select distinct FbdBatchName,FbdLotNumber, Max(date_time2) as BATCHLOT from FbpReportAuto group by FbdBatchName,FbdLotNumber order by max(date_time2) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()
        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))  
    
    def fbd_auto_rep_execute(self):
        confirm_before_export_rmg_report=mb.askquestion('Export File', 'Are you Sure you want to View the FBD Batch Report')
        if confirm_before_export_rmg_report == "yes" :
            threading.Thread(target= self.generate_fbd_report).start()
    def fbd_auto_alarm_execute(self):
        confirm_before_export_rmg_alarm=mb.askquestion('Export File', 'Are you Sure you want to View the FBD ALARM Report')
        if confirm_before_export_rmg_alarm == "yes" :
            threading.Thread(target= self.generate_fbd_alarm).start()
    def fbd_auto_audit_execute(self):
        confirm_before_export_rmg_audit=mb.askquestion('Export File', 'Are you Sure you want to View the FBD AUDIT Report')
        if confirm_before_export_rmg_audit == "yes" :
            threading.Thread(target= self.generate_fbd_audit).start()
    
    def generate_fbd_report(self):
        #commands1 = ("SELECT format(date_time2,'dd/mm/yyy    hh:mm:ss'),FbdInletTempSV,FbdInletTempAct,FbdOutletTempSV,FbdOutletTempAct,FbdBedTempSV,FbdBedTempAct,FbdInletPosSV,FbdInletValveAct,FbdInletVelocityAct,FbdBlowerRPMSV,FbdRpmAct,FbdRemark  from FbpReportAuto where (FbdBatchName = (?)) AND (FbdLotNumber = (?)) AND (FbdRemark = 'FBD HEATING PROCESS STARTED') order by date_time2") 
        #cursor.execute(commands1,batchname_fbd,lotnumber_fbd)
        #rows1=cursor.fetchall()
        #conn.commit()
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")

        commands2 = ("SELECT format(date_time2,'dd/MM/yyy    HH:mm:ss'),FbdInletTempSV,convert(numeric(10,1),FbdInletTempAct) as Total,FbdOutletTempSV,convert(numeric(10,1),FbdOutletTempAct) as Total,FbdBedTempSV,convert(numeric(10,1),FbdBedTempAct) as Total,FbdInletPosSV,FbdInletValveAct,convert(numeric(10,1),FbdInletVelocityAct) as Total,FbdBlowerRPMSV,FbdRpmAct,FbdRemark  from FbpReportAuto where (FbdBatchName = (?)) AND (FbdLotNumber = (?)) AND (FbdRemark = 'FBD HEATING PROCESS STARTED' or FbdRemark = 'FBD PROCESS RUNNING' or FbdRemark = 'FBD COOLING PROCESS STARTED' or FbdRemark = 'FBD FINAL BAG SHAKING STARTED' or FbdRemark = 'FBD PROCESS COMPLETED') order by date_time2") 
        cursor.execute(commands2,batchname_fbd,lotnumber_fbd)
        rows2=cursor.fetchall()
        conn.commit()
        batch_data = rows2 
        commands3 = ("SELECT MIN(format(date_time2,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time2,'dd/MM/yyy    HH:mm:ss')) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands3,batchname_fbd,lotnumber_fbd)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

                    ##              0       1           2           3           4           5           6                7              8                 9         10                      11                 12                    13             14              15          16              17              18                  19          20                                                                       
        commands4=("Select FbdProdName,FbdProdCode,FbdBatchName,FbdLotNumber,FbdEqpID,FbdBatchSize,FbdPrintInterval,FbdTotalTimeSV,FbdHeatingTime,FbdCoolingTime,FbdFinalBAgShakeTime,FbdBagShakeTimeSV,FbdBagShakeInterval,FbdInletTempSV,FbdInletTempTolSV,FbdBlowerRPMSV,FbdInletPosSV,FbdOutletTempSV,FbdOutletTempTolSV,FbdBedTempSV,FbdBedTempTolSV from FbpReportAuto where (FbdBatchName = (?)) and (FbdLotNumber = (?)) and (FbdRemark = 'FBD HEATING PROCESS STARTED' OR FbdRemark = 'FBD PROCESS RUNNING') order by date_time2")
        cursor.execute(commands4,batchname_fbd,lotnumber_fbd)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands_user_time_fbd = ("Select Min(date_time2) AS StartTime, MAX(date_time2) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands_user_time_fbd,batchname_fbd,lotnumber_fbd)
        rows_user=cursor.fetchall()
        conn.commit()
        sort_strt_end_time_fbd = [item for t in rows_user for item in t]
        start_time_user_fbd = sort_strt_end_time_fbd[0]
        end_time_user_fbd = sort_strt_end_time_fbd[1]

        commands5 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?) ")
        cursor.execute(commands5,batchname_fbd,lotnumber_fbd,start_time_user_fbd)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?)")
        cursor.execute(commands6,batchname_fbd,lotnumber_fbd,end_time_user_fbd)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")

        wb = load_workbook('E:\\SCADA\\Format\\FBD_Auto_Batch_Report.xlsm',read_only=False,keep_vba=True)
        ws = wb.worksheets[0]
        a = 3
        b = 15
        c=19
        d=23
        e=35
        ## batch start and end time
        c14 = ws.cell(row=11,column=2)
        c14.value = str(start_time)
        c15 = ws.cell(row=11,column=10)
        c15.value = str(end_time)
        c16 = ws.cell(row=43,column=2)
        c16.value = str(start_time)
        c17 = ws.cell(row=43,column=10)
        c17.value = str(end_time)
        ## batch start and end user
        c18 = ws.cell(row=12,column=2)
        c18.value = str(batch_start_user)
        c19 = ws.cell(row=12,column=10)
        c19.value = str(batch_end_user)
        c20 = ws.cell(row=44,column=2)
        c20.value = str(batch_start_user)
        c21 = ws.cell(row=44,column=10)
        c21.value = str(batch_end_user)
        c24 = ws.cell(row=13,column=10)
        c24.value = str(current_dt_time)
        c25 = ws.cell(row=45,column=10)
        c25.value = str(current_dt_time)
        current_user_cell = ws.cell(row=13,column=2)
        current_user_cell.value = str(current_user)
        current_user_cell2 = ws.cell(row=45,column=2)
        current_user_cell2.value = str(current_user)
        ## recipe loop
        for x in rows_sp[0:7] :
            c1 = ws.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        for x in rows_sp[7:10] :
            c2 = ws.cell(row=b+1,column=2)
            c2.value = str(x)
            b+=1
        for x in rows_sp[10:13] :
            c3 = ws.cell(row=c+1,column=2)
            c3.value = str(x)
            c+=1
        for x in rows_sp[13:21] :
            c4 = ws.cell(row=d+1,column=2)
            c4.value = str(x)
            d+=1
        ## recipe loop for second page
        for x in rows_sp[0:7] :
            c5 = ws.cell(row=e+1,column=2)
            c5.value = str(x)
            e+=1
        ## process interval insert data
        p = 48
        for i in batch_data:
            k = 0
            for j in i:
                c25 = ws.cell(row=p+1,column=k+1)
                c25.value = str(j)
                k+=1
            p+=1
        btch_filename_rmg_alarm = "trail_testing_fbd_batch"
        filename = "FBD_Auto_Batch_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\FBD\\EXEC\\"
        xl_path1 = xl_path + filename
        wb.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\FBD\\EXEC\\FBD_Auto_Batch_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb.close()
        sleep(2)
        root.destroy()
    
    ## FBD alarm fetch
    def generate_fbd_alarm(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        commands2= ("SELECT MIN(date_time2) AS StartTime, MAX(date_time2) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands2,batchname_fbd,lotnumber_fbd)
        rmg_alarm_start_end_time=cursor.fetchall()
        sort_strt_end_time = [item for t in rmg_alarm_start_end_time for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(date_time2,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time2,'dd/MM/yyy    HH:mm:ss')) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands3,batchname_fbd,lotnumber_fbd)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]
                    ##              0       1           2           3           4           5           6                
        commands4=("Select FbdProdName,FbdProdCode,FbdBatchName,FbdLotNumber,FbdEqpID,FbdBatchSize,FbdPrintInterval from FbpReportAuto where (FbdBatchName = (?)) and (FbdLotNumber = (?)) and (FbdRemark = 'FBD HEATING PROCESS STARTED' OR FbdRemark = 'FBD PROCESS RUNNING' OR FbdRemark = 'AIR PRESSURE LOW') order by date_time2")
        cursor.execute(commands4,batchname_fbd,lotnumber_fbd)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands5 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?) ")
        cursor.execute(commands5,batchname_fbd,lotnumber_fbd,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?)")
        cursor.execute(commands6,batchname_fbd,lotnumber_fbd,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        alarm_value_fetch = "Select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) and (Al_Group = 3 OR Al_Message = 'POWER RESUMED' OR Al_Message = 'SCADA COMMUNICATION ESTABLISHED' OR Al_Message = 'SCADA COMMUNICATION FAILURE' OR Al_Message = 'POWER FAILURE') order by Al_Start_Time"
        cursor.execute(alarm_value_fetch,start_time,end_time)
        alarm_values=cursor.fetchall()
        wb1 = load_workbook('E:\\SCADA\\Format\\FBD_Auto_Alarm_Report.xlsm',read_only=False,keep_vba=True)
        ws1 = wb1.worksheets[0]
        btch_filename_rmg_alarm = "testing_alarm_fbd"
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws1.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws1.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws1.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws1.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws1.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws1.cell(row=12,column=4)
        c24.value = str(current_dt_time)
        current_user_cell = ws1.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(alarm_values)==0:
            c17 = ws1.cell(row=14,column=3)
            c17.value = str("NO ALARM FOUND REPORT")
        else:
            for i in alarm_values:
                k = 0
                for j in i:
                    c25 = ws1.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "FBD_Auto_Alarm_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\FBD\\EXEC\\"
        xl_path1 = xl_path + filename
        wb1.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\FBD\\EXEC\\FBD_Auto_Alarm_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb1.close()
        sleep(2)
        root.destroy()
## FBD Audit Fetch
    def generate_fbd_audit(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        commands2= ("SELECT MIN(date_time2) AS StartTime, MAX(date_time2) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands2,batchname_fbd,lotnumber_fbd)
        rmg_alarm_start_end_time=cursor.fetchall()
        sort_strt_end_time = [item for t in rmg_alarm_start_end_time for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(date_time2,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(date_time2,'dd/MM/yyy    HH:mm:ss')) FROM FbpReportAuto WHERE FbdBatchName = (?) AND FbdLotNumber = (?)")
        cursor.execute(commands3,batchname_fbd,lotnumber_fbd)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]

                    ##              0       1           2           3           4           5           6                
        commands4=("Select FbdProdName,FbdProdCode,FbdBatchName,FbdLotNumber,FbdEqpID,FbdBatchSize,FbdPrintInterval from FbpReportAuto where (FbdBatchName = (?)) and (FbdLotNumber = (?)) and (FbdRemark = 'FBD HEATING PROCESS STARTED' OR FbdRemark = 'FBD PROCESS RUNNING') order by date_time2")
        cursor.execute(commands4,batchname_fbd,lotnumber_fbd)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands5 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?) ")
        cursor.execute(commands5,batchname_fbd,lotnumber_fbd,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select FbdUsername from FbpReportAuto where FbdBatchName = (?) and FbdLotNumber = (?) and date_time2 = (?)")
        cursor.execute(commands6,batchname_fbd,lotnumber_fbd,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        audit_value_fetch = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'RMG%') AND (Ev_Message NOT LIKE 'BLENDER%') order by Ev_Time"
        cursor.execute(audit_value_fetch,start_time,end_time)
        audit_value_fetch=cursor.fetchall()
        btch_filename_rmg_alarm = "Audit_testing1_fbd"
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        wb2 = load_workbook('E:\\SCADA\\Format\\FBD_Auto_Audit_Report.xlsm',read_only=False,keep_vba=True)
        ws2 = wb2.worksheets[0]
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws2.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws2.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws2.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws2.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws2.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws2.cell(row=12,column=4)
        c24.value = str(current_dt_time)
        current_user_cell = ws2.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(audit_value_fetch)==0:
            c17 = ws2.cell(row=14,column=2)
            c17.value = str("NO AUDIT FOUND REPORT")
        else:
            for i in audit_value_fetch:
                k = 0
                for j in i:
                    c25 = ws2.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "FBD_Auto_Audit_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\FBD\\EXEC\\"
        xl_path1 = xl_path + filename
        wb2.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\FBD\\EXEC\\FBD_Auto_Audit_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb2.close()
        sleep(2)
        root.destroy()

class bld_report_fetch():
    def __init__ (self):
        global sytem_selected
        sytem_selected = report_types_system.get()
        generate_report_rmg_auto.place_forget(),generate_report_rmg_alarm.place_forget(),generate_report_rmg_audit.place_forget()
        ent_search.place_forget(),clear_rmg_record.place_forget(),search_rmg_data.place_forget()
        generate_report_fbd_auto.place_forget(),generate_report_fbd_alarm.place_forget(),generate_report_fbd_audit.place_forget()
        ent_search_fbd.place_forget(),clear_fbd_record.place_forget(),search_fbd_data.place_forget()
        alrm_aud_btn.place_forget()
        alrm_btn.place_forget()
        ent_search.place_forget()
        clear_rmg_record.place_forget()
        search_rmg_data.place_forget()
        my_tree.delete(*my_tree.get_children())
        commands = ("select distinct BlBatchName,BlLotName, Max(Date_Time3) as BATCHLOT from BlReportAuto group by BlBatchName,BlLotName order by max(Date_Time3) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()

        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))
        def val (event):
            global select_report   
            global batchname_bld
            global lotnumber_bld
            item1 = my_tree.focus()
            item = my_tree.item(item1,option="values")
            #item = my_tree.item(my_tree.focus())
            #print(item['values'][0])
            batchname_bld1 = item[0]
            batchname_bld = str(batchname_bld1)
            lotnumber_bld1 = item[1]
            lotnumber_bld = str(lotnumber_bld1)
            generate_report_bld_auto.place(x= 650,y=70)
            generate_report_bld_alarm.place(x = 850 , y = 70)
            generate_report_bld_audit.place(x = 1050, y = 70)
            generate_report_bld_auto.config(command = self.bld_auto_rep_execute)
            generate_report_bld_alarm.config(command = self.bld_auto_alarm_execute)
            generate_report_bld_audit.config(command = self.bld_auto_audit_execute)
        #my_tree.bind('<Double-1>',val)
        my_tree.bind('<ButtonRelease-1>',val)
        search_bld_data.place(x= 1000,y=200)
        lbl_search.place(x = 650, y = 200)
        ent_search_bld.place(x = 750, y = 210)
        keyboard_btn.place(x = 650, y = 170)
        clear_bld_record.place( x= 1150, y= 200)
        alrm_btn.place(x=750, y = 250)
        alrm_aud_btn.place(x=1050, y = 250)
        alrm_btn.config(command = report_fetch)
        alrm_aud_btn.config(command = report_fetch)
        show_aud_alrm_reprt.config(command = audit_alarm_reports)
        show_aud_alrm_reprt.config(command = audit_alarm_reports)
        search_bld_data.config(command = self.search_bld_data)
        clear_bld_record.config(command = self.clear_records_bld)
    ## search BLENDER distinct value
    def search_bld_data(self):
        get_search_bld_value = ent_search_bld.get()
        querry = "select distinct BlBatchName,BlLotName as BATCHLOT from BlReportAuto where (BlBatchName = (?)) OR (BlLotName = (?)) "
        cursor.execute(querry,get_search_bld_value,get_search_bld_value)
        get_filter_vlve=cursor.fetchall()
        conn.commit()
        #print(get_filter_vlve)
        if len(get_filter_vlve)==0:
            my_tree.delete(*my_tree.get_children())
            my_tree.insert('',index='end',values = ("NO RECORD FOUND"))
        else:
            my_tree.delete(*my_tree.get_children())
            for values in get_filter_vlve:
                batchname = str(values[0])
                lotnumber = str(values[1])
                my_tree.insert('',index='end',values = (batchname,lotnumber))
    
    ## clear search BLENDER records
    def clear_records_bld(self):
        my_tree.delete(*my_tree.get_children())
        commands = ("select distinct BlBatchName,BlLotName, Max(Date_Time3) as BATCHLOT from BlReportAuto group by BlBatchName,BlLotName order by max(Date_Time3) desc")
        cursor.execute(commands)
        rows1=cursor.fetchall()
        conn.commit()
        for values in rows1:
            file_date =  str(values[2])
            file_date = file_date[:19]
            batchname = str(values[0])
            lotnumber = str(values[1])
            my_tree.insert('',index='end',values = (batchname,lotnumber,file_date))

    def bld_auto_rep_execute(self):
        confirm_before_export_rmg_report=mb.askquestion('Export File', 'Are you Sure you want to View the BLENDER Batch Report')
        if confirm_before_export_rmg_report == "yes" :
            threading.Thread(target= self.generate_bld_report).start()
    def bld_auto_alarm_execute(self):
        confirm_before_export_rmg_alarm=mb.askquestion('Export File', 'Are you Sure you want to View the BLENDER ALARM Report')
        if confirm_before_export_rmg_alarm == "yes" :
            threading.Thread(target= self.generate_bld_alarm).start()
    def bld_auto_audit_execute(self):
        confirm_before_export_rmg_audit=mb.askquestion('Export File', 'Are you Sure you want to View the BLENDER AUDIT Report')
        if confirm_before_export_rmg_audit == "yes" :
            threading.Thread(target= self.generate_bld_audit).start()
    ## blender fetch auto report
    def generate_bld_report(self):
        #print(lotnumber_bld)
        #commands1 = ("SELECT format(Date_Time3,'dd/mm/yyy    hh:mm:ss'),BlRpmSV,BlenderRpmAct,BlRemark  from BlReportAuto where (BlBatchName = (?)) AND (BlLotName = (?)) AND (BlRemark = 'BLENDER PROCESS STARTED') order by Date_Time3") 
        #cursor.execute(commands1,batchname_bld,lotnumber_bld)
        #rows1=cursor.fetchall()
        #conn.commit()
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")

        commands2 = ("SELECT format(Date_Time3,'dd/MM/yyy    HH:mm:ss'),BlRpmSV,BlenderRpmAct,BlRemark  from BlReportAuto where (BlBatchName = (?)) AND (BlLotName = (?)) AND (BlRemark = 'BLENDER PROCESS STARTED' or BlRemark = 'BLENDER PROCESS RUNNING' or BlRemark = 'BLENDER PROCESS COMPLETED') order by Date_Time3") 
        cursor.execute(commands2,batchname_bld,lotnumber_bld)
        rows2=cursor.fetchall()
        conn.commit()
        batch_data = rows2 
        commands3 = ("SELECT MIN(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands3,batchname_bld,lotnumber_bld)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]
                ##            0         1           2              3        4       5           6             7         8
        commands4=("Select BlProdName,BlProdCode,BlBatchName,BlLotName,BlBinID,BlBatchSize,BlPrintInterval,BlRpmSV,SBlendTime  from BlReportAuto where (BlBatchName = (?)) and (BlLotName = (?)) and (BlRemark = 'BLENDER PROCESS STARTED' OR BlRemark = 'BLENDER PROCESS RUNNING') order by Date_Time3")
        cursor.execute(commands4,batchname_bld,lotnumber_bld)
        rows_sp=cursor.fetchone()
        conn.commit()
        
        commands_user_time_bld = ("SELECT MIN(Date_Time3) AS StartTime, MAX(Date_Time3) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands_user_time_bld,batchname_bld,lotnumber_bld)
        rows_user=cursor.fetchall()
        conn.commit()
        sort_strt_end_time_bld = [item for t in rows_user for item in t]
        start_time_user_bld = sort_strt_end_time_bld[0]
        end_time_user_bld = sort_strt_end_time_bld[1]

        commands5 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands5,batchname_bld,lotnumber_bld,start_time_user_bld)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands6,batchname_bld,lotnumber_bld,end_time_user_bld)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        wb = load_workbook('E:\\SCADA\\Format\\BLENDER_Auto_Batch_Report.xlsm',read_only=False,keep_vba=True)
        ws = wb.worksheets[0]
        BlRpmSV =  str(rows_sp[7])
        SBlendTime =  str(rows_sp[8])
        a = 3
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        ## recipe loop
        for x in rows_sp[0:7] :
            c1 = ws.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws.cell(row=11,column=2)
        c20.value = str(start_time)
        c21 = ws.cell(row=11,column=7)
        c21.value = str(end_time)
        c22 = ws.cell(row=12,column=2)
        c22.value = str(batch_start_user)
        c23 = ws.cell(row=12,column=7)
        c23.value = str(batch_end_user)
        c24 = ws.cell(row=14,column=2)
        c24.value = str(BlRpmSV)
        c25 = ws.cell(row=14,column=7)
        c25.value = str(SBlendTime)
        c26 = ws.cell(row=13,column=7)
        c26.value = str(current_dt_time)
        current_user_cell = ws.cell(row=13,column=2)
        current_user_cell.value = str(current_user)
    ## process interval insert data
        p = 17
        for i in batch_data:
            k = 0
            for j in i:
                c25 = ws.cell(row=p+1,column=k+1)
                c25.value = str(j)
                k+=1
            p+=1
        btch_filename_rmg_alarm = "trail_testing_bld_batch"
        filename = "BLD_Auto_Batch_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\"
        xl_path1 = xl_path + filename
        wb.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\BLD_Auto_Batch_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb.close()
        sleep(2)
        root.destroy()
        ## BLENDER alarm fetch
    def generate_bld_alarm(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        commands2= ("SELECT MIN(Date_Time3) AS StartTime, MAX(Date_Time3) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands2,batchname_bld,lotnumber_bld)
        rmg_alarm_start_end_time=cursor.fetchall()
        sort_strt_end_time = [item for t in rmg_alarm_start_end_time for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands3,batchname_bld,lotnumber_bld)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]
                ##            0         1           2              3        4       5           6             7         8
        commands4=("Select BlProdName,BlProdCode,BlBatchName,BlLotName,BlBinID,BlBatchSize,BlPrintInterval,BlRpmSV,SBlendTime  from BlReportAuto where (BlBatchName = (?)) and (BlLotName = (?)) and (BlRemark = 'BLENDER PROCESS STARTED' OR BlRemark = 'BLENDER PROCESS RUNNING') order by Date_Time3")
        cursor.execute(commands4,batchname_bld,lotnumber_bld)
        rows_sp=cursor.fetchone()
        conn.commit()


        commands5 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands5,batchname_bld,lotnumber_bld,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands6,batchname_bld,lotnumber_bld,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user))

        alarm_value_fetch = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) and (Al_Group = 2 OR Al_Message = 'POWER RESUMED' OR Al_Message = 'SCADA COMMUNICATION ESTABLISHED' OR Al_Message = 'SCADA COMMUNICATION FAILURE' OR Al_Message = 'POWER FAILURE') order by Al_Start_Time"
        cursor.execute(alarm_value_fetch,start_time,end_time)
        alarm_values=cursor.fetchall()
        wb1 = load_workbook('E:\\SCADA\\Format\\BLENDER_Auto_Alarm_Report.xlsm',read_only=False,keep_vba=True)
        ws1 = wb1.worksheets[0]
        btch_filename_rmg_alarm = "testing_alarm_fbd"
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws1.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws1.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws1.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws1.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws1.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws1.cell(row=12,column=4)
        c24.value = str(current_dt_time)
        current_user_cell = ws1.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(alarm_values)==0:
            c17 = ws1.cell(row=14,column=3)
            c17.value = str("NO ALARM FOUND REPORT")
        else:
            for i in alarm_values:
                k = 0
                for j in i:
                    c25 = ws1.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "BLENDER_Auto_Alarm_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\"
        xl_path1 = xl_path + filename
        wb1.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\BLENDER_Auto_Alarm_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb1.close()
        sleep(2)
        root.destroy()
    
    def generate_bld_audit(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        commands2= ("SELECT MIN(Date_Time3) AS StartTime, MAX(Date_Time3) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands2,batchname_bld,lotnumber_bld)
        rmg_alarm_start_end_time=cursor.fetchall()
        sort_strt_end_time = [item for t in rmg_alarm_start_end_time for item in t]
        start_time = sort_strt_end_time[0]
        end_time = sort_strt_end_time[1]

        commands3 = ("SELECT MIN(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) AS StartTime, MAX(format(Date_Time3,'dd/MM/yyy    HH:mm:ss')) FROM BlReportAuto WHERE BlBatchName = (?) AND BlLotName = (?)")
        cursor.execute(commands3,batchname_bld,lotnumber_bld)
        rows3=cursor.fetchall()
        conn.commit()
        sort_strt_end_time = [item for t in rows3 for item in t]
        start_time_report = sort_strt_end_time[0]
        end_time_report = sort_strt_end_time[1]
                ##            0         1           2              3        4       5           6             7         8
        commands4=("Select BlProdName,BlProdCode,BlBatchName,BlLotName,BlBinID,BlBatchSize,BlPrintInterval,BlRpmSV,SBlendTime  from BlReportAuto where (BlBatchName = (?)) and (BlLotName = (?)) and (BlRemark = 'BLENDER PROCESS STARTED' OR BlRemark = 'BLENDER PROCESS RUNNING') order by Date_Time3")
        cursor.execute(commands4,batchname_bld,lotnumber_bld)
        rows_sp=cursor.fetchone()
        conn.commit()

        commands5 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands5,batchname_bld,lotnumber_bld,start_time)
        batch_start_user=cursor.fetchone()
        conn.commit()

        commands6 = ("Select BlUserName from BlReportAuto where BlBatchName = (?) and BlLotName = (?) and Date_Time3 = (?)")
        cursor.execute(commands6,batchname_bld,lotnumber_bld,end_time)
        batch_end_user=cursor.fetchone()
        conn.commit()

        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user)) 

        audit_value_fetch = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'RMG%') AND (Ev_Message NOT LIKE 'FBD%') order by Ev_Time"
        cursor.execute(audit_value_fetch,start_time,end_time)
        audit_value_fetch=cursor.fetchall()
        btch_filename_rmg_alarm = "Audit_testing1_BLENDER"
        wb2 = load_workbook('E:\\SCADA\\Format\\BLENDER_Auto_Audit_Report.xlsm',read_only=False,keep_vba=True)
        ws2 = wb2.worksheets[0]
        batch_start_user =  str(batch_start_user[0])
        batch_end_user =  str(batch_end_user[0])
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        ## recipe loop 
        a = 2
        for x in rows_sp[0:7] :
            c1 = ws2.cell(row=a+1,column=2)
            c1.value = str(x)
            a+=1
        c20 = ws2.cell(row=10,column=2)
        c20.value = str(start_time_report)
        c21 = ws2.cell(row=10,column=4)
        c21.value = str(end_time_report)
        c22 = ws2.cell(row=11,column=2)
        c22.value = str(batch_start_user)
        c23 = ws2.cell(row=11,column=4)
        c23.value = str(batch_end_user)
        c24 = ws2.cell(row=12,column=4)
        c24.value = str(current_dt_time)      
        current_user_cell = ws2.cell(row=12,column=2)
        current_user_cell.value = str(current_user)
        p = 13
        if len(audit_value_fetch)==0:
            c17 = ws2.cell(row=14,column=2)
            c17.value = str("NO AUDIT FOUND REPORT")
        else:
            for i in audit_value_fetch:
                k = 0
                for j in i:
                    c25 = ws2.cell(row=p+1,column=k+1)
                    c25.value = str(j)
                    k+=1
                p+=1
        sleep(2)
        filename = "BLENDER_Auto_Alarm_Report_exe" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\"
        xl_path1 = xl_path + filename
        wb2.save(xl_path1)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\BLENDER\\EXEC\\BLENDER_Auto_Alarm_Report_exe.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename_rmg_alarm)
        excel1.Quit()
        wb2.close()
        sleep(2)
        root.destroy()

class Timewise_alarm_audit_system_report():
    def __init__ (self):        
        global sytem_selected
        sytem_selected = report_types_system.get()
        print(sytem_selected)
        my_tree.delete(*my_tree.get_children())
        generate_report_rmg_auto.place_forget(),generate_report_rmg_alarm.place_forget(),generate_report_rmg_audit.place_forget()
        search_rmg_data.place_forget(),lbl_search.place_forget,ent_search.place_forget(),keyboard_btn.place_forget(),clear_rmg_record.place_forget(),alrm_btn.place_forget(),alrm_aud_btn.place_forget(),show_aud_alrm_reprt.place_forget()
        generate_report_fbd_auto.place_forget(),generate_report_fbd_alarm.place_forget(),generate_report_fbd_audit.place_forget()
        ent_search_fbd.place_forget(),clear_fbd_record.place_forget(),search_fbd_data.place_forget()
        generate_report_bld_auto.place_forget(),generate_report_bld_alarm.place_forget(),generate_report_bld_audit.place_forget()
        ent_search_bld.place_forget(),clear_bld_record.place_forget(),search_bld_data.place_forget()
        alrm_btn.place(x=750, y = 200)
        alrm_aud_btn.place(x=1050, y = 200)
        alrm_btn.config(command = report_fetch)
        alrm_aud_btn.config(command = report_fetch)
        show_aud_alrm_reprt.config(command = audit_alarm_reports)
class audit_alarm_reports():
    def __init__ (self):
        global get_alarm_values
        global get_alarm_audit_values
        global path_alarm_audit_timewise
        global btch_filename
        global path
        #print(sytem_selected)
        #print(alarm_audit_selected)
        ## RMG ALARM AND AUDIT FETCH
        if sytem_selected == "RMG" and alarm_audit_selected == "ALARM" :
            get_alarm_command = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) and Al_Group = 1 order by Al_Start_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_alarm_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "RMG_Timewise_Alarm_Report_FILTER"
            filename_rmg = "RMG_Timewise_Alarm_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        
        if sytem_selected == "RMG" and alarm_audit_selected == "AUDIT" :
            get_audit_command = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'FBD%') AND (Ev_Message NOT LIKE 'BLENDER%') order by Ev_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_audit_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "RMG_Timewise_Audit_Report_FILTER"
            filename_rmg = "RMG_Timewise_Audit_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        ## FBD ALARM AND AUDIT FETCH
        if sytem_selected == "FBD" and alarm_audit_selected == "ALARM" :
            get_alarm_command = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) AND Al_Start_Time <= (?) AND (Al_Group = 3 OR Al_Message = 'POWER RESUMED' OR Al_Message = 'SCADA COMMUNICATION ESTABLISHED' OR Al_Message = 'SCADA COMMUNICATION FAILURE' OR Al_Message = 'POWER FAILURE' OR Al_Message = 'AIR PRESSURE LOW') order by Al_Start_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_alarm_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "FBD_Timewise_Alarm_Report_FILTER"
            filename_rmg = "FBD_Timewise_Alarm_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        
        if sytem_selected == "FBD" and alarm_audit_selected == "AUDIT" :
            get_audit_command = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'RMG%') AND (Ev_Message NOT LIKE 'BLENDER%') order by Ev_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_audit_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "FBD_Timewise_Audit_Report_FILTER"
            filename_rmg = "FBD_Timewise_Audit_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        
        ## BLENDER ALARM AND AUDIT FETCH
        if sytem_selected == "BLENDER" and alarm_audit_selected == "ALARM" :
            get_alarm_command = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) and (Al_Group = 2 OR Al_Message = 'POWER RESUMED' OR Al_Message = 'SCADA COMMUNICATION ESTABLISHED' OR Al_Message = 'SCADA COMMUNICATION FAILURE' OR Al_Message = 'POWER FAILURE') order by Al_Start_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_alarm_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "BLENDER_Timewise_Alarm_Report_FILTER"
            filename_rmg = "BLENDER_Timewise_Alarm_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        
        if sytem_selected == "BLENDER" and alarm_audit_selected == "AUDIT" :
            get_audit_command = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) AND (Ev_Message NOT LIKE 'RMG%') AND (Ev_Message NOT LIKE 'FBD%') order by Ev_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_audit_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "BLENDER_Timewise_Audit_Report_FILTER"
            filename_rmg = "BLENDER_Timewise_Audit_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
    ## TIMEWISE ALL SYSTEM ALARM FETCH
        if sytem_selected == "TIMEWISEFILTER" and alarm_audit_selected == "ALARM" :
            get_alarm_command = "select format(Al_Event_Time,'dd/MM/yyy    HH:mm:ss'),format(Al_Norm_Time,'dd/MM/yyy    HH:mm:ss'),Al_Message,Al_User from ALARMHISTORY where Al_Start_Time >= (?) and Al_Start_Time <= (?) order by Al_Start_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_alarm_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "ALL_SYSTEM_Timewise_Alarm_Report_FILTER"
            filename_rmg = "ALL_SYSTEM_Timewise_Alarm_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        
        if sytem_selected == "TIMEWISEFILTER" and alarm_audit_selected == "AUDIT" :
            get_audit_command = "select format(Ev_Time,'dd/MM/yyy    HH:mm:ss'),Ev_Message,Ev_Prev_Value,Ev_Value,Ev_User from EVENTHISTORY where Ev_Time >= (?) and Ev_Time <= (?) order by Ev_Time"
            args_strt_end_time = (from_date_time, to_date_time)
            cursor.execute(get_audit_command,args_strt_end_time)
            get_alarm_audit_values = cursor.fetchall()
            btch_filename = "ALL_SYSTEM_Timewise_Audit_Report_FILTER"
            filename_rmg = "ALL_SYSTEM_Timewise_Audit_Report" + ".xlsm"
            xl_path_rmg = "E:\\SCADA\\Format\\"
            path = xl_path_rmg + filename_rmg
            self.get_alarm_audit_report()
        

    def get_alarm_audit_report(self):
        try:
            os.system('TASKKILL /F /IM AcroRd32.exe')

        except Exception:
            print("KU")
        wb10 = load_workbook(path,read_only=False,keep_vba=True)
        ws10 = wb10.worksheets[0]
        c25 = ws10.cell(row=3,column=2)
        c25.value = str(from_date_time1)
        c25 = ws10.cell(row=3,column=4)
        c25.value = str(to_date_time1)
        current_time = datetime.now()
        current_dt_time = current_time.strftime(" %d/%m/%Y  %H:%M:%S")
        c24 = ws10.cell(row=4,column=4)
        c24.value = str(current_dt_time)
        commands7 = ("Select USERNAME AS username from UN")
        cursor.execute(commands7)
        current_user = cursor.fetchall()
        conn.commit()
        current_user = (current_user[0])
        current_user = (" ".join(current_user))
        
        current_user_cell = ws10.cell(row=4,column=2)
        current_user_cell.value = str(current_user)
        if len(get_alarm_audit_values) == 0: 
            c17 = ws10.cell(row=7,column=3)
            c17.value = str("NO RECORD FOUND")
        else : 
            p = 5
            for i in get_alarm_audit_values:
                k = 0
                for j in i:
                    c1 = ws10.cell(row=p+1,column=k+1)
                    c1.value = str(j)
                    k+=1
                p+=1
        sleep(1)
        
        filename = "Audit_Alarm_timewise_report" + ".xlsm"
        xl_path = "E:\\SCADA\\File_Gen\\Timewise_Filter\\EXEC\\"
        xl_path1 = xl_path + filename
        wb10.save(xl_path1)
        sleep(3)
        excel1 = client.DispatchEx("Excel.Application")
        excel1.Application.Run("'E:\\SCADA\\File_Gen\\Timewise_Filter\\EXEC\\Audit_Alarm_timewise_report.xlsm'!Module1.SaveActiveSheetsAsPDF",btch_filename)
        excel1.Quit()
        wb10.close()
        sleep(2)
        root.destroy()

class report_fetch():
    def __init__ (self):
        global report_fetch_gui
        report_fetch_gui = Toplevel(root)
        report_fetch_gui.geometry("750x300")
        #report_fetch_gui.title("Select Date and Time")
        x_position = 710
        y_position = 200
        report_fetch_gui.geometry(f"750x300+{x_position}+{y_position}")
        global alarm_audit_selected
        alarm_audit_selected = alarm_audit_types.get()
        from_hr_selc = IntVar(report_fetch_gui)
        from_min_selc = IntVar(report_fetch_gui)
        from_sec_selc = IntVar(report_fetch_gui) 
        to_hr_selc = IntVar(report_fetch_gui)
        to_min_selc = IntVar(report_fetch_gui)
        to_sec_selc = IntVar(report_fetch_gui)
        #HighlightFont_fetch = Font (family='Helvetica', size=5, weight='bold')
    
        Label(report_fetch_gui,text = "Choose From Date" ,font = ("Arial", 15)).place(x = 10, y = 10)
        selec_date = DateEntry(report_fetch_gui, font = "Arial 14", selectmode = 'day')
        selec_date.place(x = 10, y = 50)

        Label(report_fetch_gui,text = "Choose From Time" ,font = ("Arial", 15)).place(x = 10, y = 90)
    
        Label (report_fetch_gui,text = "Hours",font = ("Arial", 10)).place (x=30, y = 120)
        hour_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24"),textvariable = from_hr_selc,width=2,font=('verdana',30))
        hour_spin.place(x = 10, y = 150)

        Label (report_fetch_gui,text = "Minutes",font = ("Arial", 10)).place (x=120, y = 120)
        min_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = from_min_selc,width=2,font=('verdana',30))
        min_spin.place(x = 110, y = 150)


        Label (report_fetch_gui,text = "Seconds",font = ("Arial", 10)).place (x=220, y = 120)
        sec_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = from_sec_selc,width=2,font=('verdana',30))
        sec_spin.place(x = 210, y = 150)
        ## to date and time selection
        Label(report_fetch_gui,text = "Choose To Date" ,font = ("Arial", 15)).place(x = 400, y = 10)
        to_selec_date = DateEntry(report_fetch_gui, font = "Arial 14", selectmode = 'day')
        to_selec_date.place(x = 400, y = 50)

        Label(report_fetch_gui,text = "Choose To Time" ,font = ("Arial", 15)).place(x = 420, y = 90)
    
        Label (report_fetch_gui,text = "Hours",font = ("Arial", 10)).place (x=430, y = 120)
        to_hour_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24"),textvariable = to_hr_selc,width=2,font=('verdana',30))
        to_hour_spin.place(x = 410, y = 150)

        Label (report_fetch_gui,text = "Minutes",font = ("Arial", 10)).place (x=520, y = 120)
        to_min_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = to_min_selc,width=2,font=('verdana',30))
        to_min_spin.place(x = 510, y = 150)


        Label (report_fetch_gui,text = "Seconds",font = ("Arial", 10)).place (x=610, y = 120)
        to_sec_spin = Spinbox(report_fetch_gui, values = ("00","01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20","21","22","23","24","25","26","27","28","29","30","31","32","33","34","35","36","37","38","39","40","41","42","43","44","45","46","47","48","49","50","51","52","53","54","55","56","57","58","59","60"), textvariable = to_sec_selc,width=2,font=('verdana',30))
        to_sec_spin.place(x = 610, y = 150)

            
        def date_selc():
            ## get from time and date 
            hr_get = hour_spin.get()
            min_get = min_spin.get()
            sec_get = sec_spin.get()
            hours = 25
            minutes = 61
            seconds = 61
            global from_value
            try :
                hr_get = int(hr_get) 
            
                min_get = int(min_get)
                sec_get = int(sec_get)
                if (hours > hr_get) and (minutes > min_get) and (seconds > sec_get):
                    print("success")
                else :
                    showerror ("Error", "Invalid Entry")

            except Exception:
                showerror ("Error", "Invalid Entry")
            else :
                str_hr = hour_spin.get()
                str_min = min_spin.get()
                str_sec =  sec_spin.get()
                date_value = selec_date.get_date()
                date_selc = date_value.strftime("%d-%m-%Y")
                from_value = str(date_selc) + " " + str(str_hr) + ":" + str(str_min) + ":" + str(str_sec)
                Label(report_fetch_gui, text = "You have selected from date&time  :  " + str(date_selc) + "/ "+ str(str_hr) + ": " + str(str_min)+ ": "+ str(str_sec)).place (x=10, y= 210)
            ## get to date and time
            to_hr_get = to_hour_spin.get()
            to_min_get = to_min_spin.get()
            to_sec_get = to_sec_spin.get()
            hours = 25
            minutes = 61
            seconds = 61
            try :
                to_hr_get = int(to_hr_get) 
            
                to_min_get = int(to_min_get)
                to_sec_get = int(to_sec_get)
                if (hours > to_hr_get) and (minutes > to_min_get) and (seconds > to_sec_get):
                    print("success")
                else :
                    showerror ("Error", "Invalid Entry")

            except Exception:
                showerror ("Error", "Invalid Entry")
            else :
                to_str_hr = to_hour_spin.get()
                to_str_min = to_min_spin.get()
                to_str_sec =  to_sec_spin.get()
                to_date_value = to_selec_date.get_date()
                #print(to_date_value)
                to_date_selc = to_date_value.strftime("%d-%m-%Y")
                to_value = str(to_date_selc) + " " + str(to_str_hr) + ":" + str(to_str_min) + ":" + str(to_str_sec)
                Label(report_fetch_gui, text = "You have selected To date&time  :  " + str(to_date_selc) + "/ "+ str(to_str_hr) + ": " + str(to_str_min)+ ": "+ str(to_str_sec)).place (x=400, y= 210)
                #print(to_value)
            global from_date_time
            global to_date_time
            global from_date_time1
            global to_date_time1
            from_date_time = datetime.strptime(from_value,'%d-%m-%Y %H:%M:%S')
            to_date_time = datetime.strptime(to_value,'%d-%m-%Y %H:%M:%S')
            from_date_time1 = from_date_time.strftime("%d/%m/%Y %H:%M:%S")
            to_date_time1 = to_date_time.strftime('%d/%m/%Y %H:%M:%S')
            try :
                if (from_date_time >= to_date_time):
                    #print("error")
                    report_fetch_gui.destroy()
                    raise showerror("Error", "TO DATE SHOULD BE GREATER THEN FROM DATE")
                    
            except :
                print("error")
                #showerror("Error", "TO DATE SHOULD BE GREATER THEN FROM DATE")
            else:
                show_aud_alrm_reprt.place(x = 900, y = 350)
                
                report_fetch_gui.destroy()

        Button(report_fetch_gui, text = "Submit",font = ("Arial", 20),command =  date_selc).place(x = 270, y = 240)


def callback():
    # so the touch keyboard is called tabtip.exe and its located in C:\Program Files\Common Files\microsoft shared\ink
    # here we run it after focus
    #os.system("C:\\Program Files\\Common Files\\microsoft shared\\ink\\tabtip.exe")
    #os.system('C:\\Windows\\system32\\osk.exe')
    os.system ("E:\SCADA\\kill_tabtip.lnk")
    sleep(1)
    os.system("C:\\PROGRA~1\\COMMON~1\\MICROS~1\\ink\\tabtip.exe")

def callback_entry(event):
    os.system ("E:\SCADA\\kill_tabtip.lnk")
    sleep(1)
    os.system("C:\\PROGRA~1\\COMMON~1\\MICROS~1\\ink\\tabtip.exe")


def main1():
    global report_type 
    global HighlightFont
    global report_types_system
    global generate_report_rmg_auto,generate_report_rmg_alarm,generate_report_rmg_audit
    global search_rmg_data,lbl_search,ent_search,keyboard_btn,clear_rmg_record
    global filter_rmg_data
    global alarm_audit_types
    global alrm_btn,alrm_aud_btn,show_aud_alrm_reprt
    global generate_report_fbd_auto,generate_report_fbd_alarm,generate_report_fbd_audit
    global ent_search_fbd,clear_fbd_record,search_fbd_data
    global filter_fbd_data
    global generate_report_bld_auto,generate_report_bld_alarm,generate_report_bld_audit
    global ent_search_bld,clear_bld_record,search_bld_data
    global filter_bld_data
    report_types_system = StringVar(root)
    filter_rmg_data = StringVar(root)
    alarm_audit_types = StringVar(root)
    filter_fbd_data = StringVar(root)
    filter_bld_data = StringVar(root)

    HighlightFont = font.Font(family='Helvetica', size=12, weight='bold')
    report_type = Radiobutton(root, text = "RMG",font= HighlightFont,height=2, width=15,indicatoron=0,variable = report_types_system, value = "RMG",command = rmg_report_fetch)
    report_type.place (x=20, y = 550)
    report_type = Radiobutton(root, text = "FBD", font= HighlightFont,height=2, width=15,indicatoron=0,variable = report_types_system, value ="FBD",command = fbd_report_fetch)
    report_type.place (x=230, y = 550)
    report_type = Radiobutton(root, text = "BLENDER",font= HighlightFont,height=2, width=15,indicatoron=0, variable = report_types_system, value ="BLENDER",command = bld_report_fetch)
    report_type.place (x=430, y = 550)
    report_type = Radiobutton(root, text = "TIMEWISE FILTER",font= HighlightFont,height=2, width=15,indicatoron=0, variable = report_types_system, value ="TIMEWISEFILTER",command = Timewise_alarm_audit_system_report)
    report_type.place (x=630, y = 550)

    generate_report_rmg_auto = Button(root, text = "VIEW REPORT RMG", font= HighlightFont,height=2, width=15)
    generate_report_rmg_alarm = Button(root, text = "VIEW ALARM RMG", font= HighlightFont,height=2, width=15)
    generate_report_rmg_audit = Button(root, text = "VIEW AUDIT RMG", font= HighlightFont,height=2, width=15)

    search_rmg_data = Button(root, text = "Search", font= HighlightFont,height=1, width=10)
    lbl_search = Label(root,text = "Search",font= HighlightFont,height=2, width=10)
    ent_search = tk.Entry(root,textvariable = filter_rmg_data,font=('calibre',10,'normal'),width = 30)
    keyboard_btn = Button(root,text = "Touch Keyboard",command = callback)

    ent_search.bind("<FocusIn>", callback_entry)
    clear_rmg_record = Button(root, text = "Clear", font= HighlightFont,height=1, width=10)
    alrm_btn =Radiobutton(root,text = "ALARM FILTER",font= HighlightFont,height=2, width=15,indicatoron=0,variable = alarm_audit_types, value = "ALARM")
    alrm_aud_btn =Radiobutton(root,text = "AUDIT FILTER",font= HighlightFont,height=2, width=15,indicatoron=0,variable = alarm_audit_types, value = "AUDIT")
    show_aud_alrm_reprt = Button(root, text = "View Report",font= HighlightFont,height=2, width=15)
    ## FBD mimic design start from here
    generate_report_fbd_auto = Button(root, text = "VIEW REPORT FBD", font= HighlightFont,height=2, width=15)
    generate_report_fbd_alarm = Button(root, text = "VIEW ALARM FBD", font= HighlightFont,height=2, width=15)
    generate_report_fbd_audit = Button(root, text = "VIEW AUDIT FBD", font= HighlightFont,height=2, width=15)
    ent_search_fbd = tk.Entry(root,textvariable = filter_fbd_data,font=('calibre',10,'normal'),width = 30)
    ent_search_fbd.bind("<FocusIn>", callback_entry)
    clear_fbd_record = Button(root, text = "Clear", font= HighlightFont,height=1, width=10)
    search_fbd_data = Button(root, text = "Search", font= HighlightFont,height=1, width=10)
    ## BLENDER mimic design start from here
    generate_report_bld_auto = Button(root, text = "VIEW REPORT BLD", font= HighlightFont,height=2, width=15)
    generate_report_bld_alarm = Button(root, text = "VIEW ALARM BLD", font= HighlightFont,height=2, width=15)
    generate_report_bld_audit = Button(root, text = "VIEW AUDIT BLD", font= HighlightFont,height=2, width=15)
    ent_search_bld = tk.Entry(root,textvariable = filter_bld_data,font=('calibre',10,'normal'),width = 30)
    ent_search_bld.bind("<FocusIn>", callback_entry)
    clear_bld_record = Button(root, text = "Clear", font= HighlightFont,height=1, width=10)
    search_bld_data = Button(root, text = "Search", font= HighlightFont,height=1, width=10)
    root.mainloop()


if(__name__=="__main__"): 
    main1()